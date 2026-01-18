import re
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font


def parse_address_sets(config_text):
    """解析所有 ip address-set 数据，返回名称到 IP 条目的映射"""
    address_sets = {}
    pattern = re.compile(
        r'ip address-set\s+"?([^"]+?)"?\s+type\s+object\s*\n'
        r'((?:\s+(?:address|description).+?\n)+)',
        re.DOTALL | re.MULTILINE
    )
    for match in pattern.finditer(config_text):
        name = match.group(1).strip()
        entries = []
        for line in match.group(2).split('\n'):
            line = line.strip()
            if line.startswith('address'):
                if 'mask' in line:
                    ip_mask = re.search(r'address \d+ (\S+ mask \d+)', line)
                    if ip_mask:
                        ip_mask_str = re.sub(r'mask (\d+)', r'/\1', ip_mask.group(1))
                        entries.append(ip_mask_str)
                elif 'range' in line:
                    ip_range = re.search(r'address \d+ range (\S+) (\S+)', line)
                    if ip_range:
                        entries.append(f"{ip_range.group(1)}-{ip_range.group(2)}")
        if entries:
            address_sets[name] = entries
    return address_sets


def parse_service_protocols(config_text):
    """解析所有服务数据，支持多种格式，包括包含 description 和 type group 的服务配置"""
    service_info = {}

    # 改进格式1：支持包含 description 的 ip service-set 定义
    pattern1 = re.compile(
        r'ip\s+service-set\s+(?:"((?:[^"]|\\")+)"|(\S+))(?:\s+type\s+object)?\s*\n'
        r'(?:\s+description\s+.+?\n)?'  # 匹配可选的 description 行
        r'((?:\s+service\s+\d+\s+.*?\n)+)',
        re.DOTALL | re.IGNORECASE
    )
    for match in pattern1.finditer(config_text):
        service_name = (match.group(1) or match.group(2)).strip()
        service_content = match.group(3)
        protocols = {}
        service_entries = re.findall(
            r'destination-port\s+(\d+)(?:\s+to\s+(\d+))?',
            service_content,
            re.IGNORECASE
        )
        proto_match = re.search(r'protocol\s+(\w+)', service_content, re.IGNORECASE)
        proto = proto_match.group(1).lower() if proto_match else 'tcp'
        for port1, port2 in service_entries:
            port = f"{port1}-{port2}" if port2 else port1
            protocols.setdefault(proto, []).append(port)
        if protocols:
            service_info[service_name] = protocols
        print(f"解析服务集: {service_name}, 协议信息: {protocols}")

    # 新增格式7：解析 type group 的服务配置
    pattern7 = re.compile(
        r'ip\s+service-set\s+(?:"((?:[^"]|\\")+)"|(\S+))\s+type\s+group\s*\n'
        r'((?:\s+service\s+\d+\s+service-set\s+\S+\s*\n)+)',
        re.DOTALL | re.IGNORECASE
    )
    for match in pattern7.finditer(config_text):
        service_name = (match.group(1) or match.group(2)).strip()
        service_content = match.group(3)
        nested_services = re.findall(r'service \d+ service-set (\S+)', service_content, re.IGNORECASE)
        if nested_services:
            service_info[service_name] = {'nested_services': nested_services}
        print(f"解析服务集: {service_name}, 嵌套服务: {nested_services}")

    # 格式2：服务名称+协议/端口
    pattern2 = re.compile(
        r'^(\S+)\s*\n(tcp|udp)\/(\d+)\s*\n',
        re.MULTILINE | re.IGNORECASE
    )
    for match in pattern2.finditer(config_text):
        service_name = match.group(1).strip()
        proto = match.group(2).lower()
        port = match.group(3)
        if service_name not in service_info:
            service_info[service_name] = {}
        service_info[service_name].setdefault(proto, []).append(port)
        print(f"解析服务: {service_name}, 协议: {proto}, 端口: {port}")

    # 格式3：动态端口描述
    pattern3 = re.compile(
        r'^(\S+)\s*\n.*?destination Port:(\d+)\b',
        re.MULTILINE | re.DOTALL
    )
    for match in pattern3.finditer(config_text):
        service_name = match.group(1).strip()
        port = match.group(2)
        proto = 'tcp'
        if service_name not in service_info:
            service_info[service_name] = {}
        service_info[service_name].setdefault(proto, []).append(port)
        print(f"解析服务: {service_name}, 协议: {proto}, 端口: {port}")

    # 格式4：独立协议声明
    pattern4 = re.compile(
        r'^(\S+)\s*\n.*?protocol\s+(\w+)\b',
        re.MULTILINE | re.IGNORECASE
    )
    for match in pattern4.finditer(config_text):
        service_name = match.group(1).strip()
        proto = match.group(2).lower()
        if service_name not in service_info:
            service_info[service_name] = {}
        service_info[service_name].setdefault(proto, [])
        print(f"解析服务: {service_name}, 协议: {proto}")

    # 格式5：多行服务定义
    pattern5 = re.compile(
        r'ip service-set\s+(\S+)\s+type\s+object\s*\n'
        r'((?:\s+service\s+\d+\s+protocol\s+\w+\s+.*?destination-port\s+\d+\s*\n)+)',
        re.DOTALL | re.IGNORECASE
    )
    for match in pattern5.finditer(config_text):
        service_name = match.group(1)
        service_content = match.group(2)
        service_matches = re.findall(
            r'protocol\s+(\w+).*?destination-port\s+(\d+)',
            service_content,
            re.IGNORECASE
        )
        if service_name not in service_info:
            service_info[service_name] = {}
        for proto, port in service_matches:
            proto = proto.lower()
            service_info[service_name].setdefault(proto, []).append(port)
        print(f"解析服务集: {service_name}, 协议信息: {service_info[service_name]}")

    # 格式6：图片中的服务配置
    pattern6 = re.compile(
        r'service\s+\d+\s+protocol\s+(\w+)\s+.*?destination-port\s+(\d+)',
        re.IGNORECASE
    )
    for match in pattern6.finditer(config_text):
        proto = match.group(1).lower()
        port = match.group(2)
        # 使用更可靠的匹配方式来获取服务名称
        service_set_match = re.search(r'ip service-set\s+(\S+)\s+type\s+object', config_text)
        if service_set_match:
            service_name = service_set_match.group(1).strip()
            if service_name not in service_info:
                service_info[service_name] = {}
            service_info[service_name].setdefault(proto, []).append(port)
        print(f"解析服务: {service_name}, 协议: {proto}, 端口: {port}")

    return service_info


def extract_rule_info(config_text, address_sets, service_info):
    """提取规则信息并关联协议端口"""
    rule_pattern = re.compile(
        r'rule name\s+(?:"([^"]+)"|(\S+))(.*?)(?=\n\s*(?:rule|#|$))',
        re.DOTALL
    )
    rules = []

    for match in rule_pattern.finditer(config_text):
        rule_name = match.group(1) or match.group(2)
        rule_content = match.group(3)

        # 提取服务名称
        # 提取服务名称（改进匹配逻辑）
        service_matches = re.findall(
            r'service\s+(?:"([^"]+)"|(?!protocol)(\S+))',  # 排除protocol关键字
            rule_content
        )
        service_names = [g1 or g2 for g1, g2 in service_matches if g1 or g2]
        unique_services = list(dict.fromkeys([n.strip() for n in service_names]))

        # 初始化协议和端口列表
        protocol_list = []
        port_list = []

        # 提取 service protocol udp destination-port 18514 格式的服务信息
        service_protocol_matches = re.findall(r'service protocol (\w+) destination-port (\d+)', rule_content, re.IGNORECASE)
        for proto, port in service_protocol_matches:
            protocol_list.append(proto)
            port_list.append(port)

        # 提取 service protocol icmp 格式的服务信息
        icmp_matches = re.findall(r'service protocol (\w+)', rule_content, re.IGNORECASE)
        for proto in icmp_matches:
            protocol_list.append(proto)

        for service in unique_services:
            if service.lower() == 'icmp':
                protocol_list.append('icmp')
                continue

            if service in service_info:
                if 'nested_services' in service_info[service]:
                    for nested_service in service_info[service]['nested_services']:
                        if nested_service in service_info:
                            for proto, ports in service_info[nested_service].items():
                                protocol_list.extend(proto for _ in ports)
                                port_list.extend(ports)
                else:
                    for proto, ports in service_info[service].items():
                        protocol_list.extend(proto for _ in ports)
                        port_list.extend(ports)
            else:
                protocol_list.append(service)

        # 去重处理
        protocol_list = list(dict.fromkeys(protocol_list))
        port_list = list(dict.fromkeys(port_list))

        # 处理地址信息
        dst_network_objects = re.findall(r'destination-address\s+address-set\s+(".*?"|\S+)', rule_content)
        unique_dst = list({d.strip('"').strip() for d in dst_network_objects})
        dst_ips = []
        for name in unique_dst:
            dst_ips.extend(address_sets.get(name, []))

        # 源地址处理
        src_addrs = re.findall(r'source-address\s+address-set\s+(".*?"|\S+)', rule_content)
        unique_src = list({a.strip('"').strip() for a in src_addrs})
        ip_entries = []
        for addr_set in unique_src:
            ip_entries.extend(address_sets.get(addr_set, []))

        rules.append({
            'name': rule_name,
            'source_zones': list(dict.fromkeys(re.findall(r'source-zone (\S+)', rule_content))),
            'destination_zones': list(dict.fromkeys(re.findall(r'destination-zone (\S+)', rule_content))),
            'source_addrs': unique_src,
            'dst_network_objects': unique_dst,
            'dst_ips': dst_ips,
            'ip_entries': ip_entries,
            'protocols': protocol_list,
            'ports': port_list
        })

        print(f"处理规则: {rule_name}")
        print(f"  源区域: {rules[-1]['source_zones']}")
        print(f"  目的区域: {rules[-1]['destination_zones']}")
        print(f"  源地址: {rules[-1]['source_addrs']}")
        print(f"  目的地址: {rules[-1]['dst_network_objects']}")
        print(f"  协议: {rules[-1]['protocols']}")
        print(f"  端口: {rules[-1]['ports']}\n")

    return rules


def write_to_excel(rules, template_path, output_path):
    """写入Excel模板，含Port列"""
    wb = load_workbook(template_path)
    ws = wb.active
    black_font = Font(color='000000')

    # 设置列标题
    headers = [
        "Rule Name", "Source Zone", "Source Address",
        "Source IP", "Destination Zone", "destination-address address-set",
        "ip address-set", "Protocol", "Port"
    ]
    for col, title in enumerate(headers, 1):
        if ws.cell(1, col).value != title:
            ws.cell(1, col, title)

    # 设置列宽
    column_widths = {
        'A': 25, 'B': 15, 'C': 25,
        'D': 40, 'E': 15, 'F': 30,
        'G': 30, 'H': 15, 'I': 20
    }
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    # 写入数据
    for row_idx, rule in enumerate(rules, 2):
        # 格式化 IP 地址和子网掩码，去掉 "/32" 前的空格
        formatted_ip_entries = [re.sub(r'(\d+)\s*/(\d+)', r'\1/\2', ip) for ip in rule['ip_entries']]
        formatted_dst_ips = [re.sub(r'(\d+)\s*/(\d+)', r'\1/\2', ip) for ip in rule['dst_ips']]

        cells = [
            rule['name'],
            '\n'.join(rule['source_zones']),
            '\n'.join(rule['source_addrs']),
            '\n'.join(formatted_ip_entries),
            '\n'.join(rule['destination_zones']),
            '\n'.join(rule['dst_network_objects']),
            '\n'.join(formatted_dst_ips),
            '\n'.join(rule['protocols']),
            '\n'.join(map(str, rule['ports']))
        ]

        for col_idx, value in enumerate(cells, 1):
            cell = ws.cell(row_idx, col_idx, value)
            cell.alignment = Alignment(wrap_text=True)
            cell.font = black_font

    wb.save(output_path)


def main():
    config_path = 'huawei_firewall.txt'
    template_path = '防火墙策略表-新加坡.xlsx'
    output_path = 'firewall_rules_report.xlsx'

    with open(config_path, 'r', encoding='utf-8') as f:
        config_text = f.read()

    print("正在解析地址集合...")
    address_sets = parse_address_sets(config_text)

    print("正在解析服务协议...")
    service_info = parse_service_protocols(config_text)

    print("正在提取规则信息...")
    rules = extract_rule_info(config_text, address_sets, service_info)

    print("正在生成Excel报告...")
    write_to_excel(rules, template_path, output_path)

    print(f"\n成功处理 {len(rules)} 条规则，报告已保存至：{output_path}")


if __name__ == "__main__":
    main()
