import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font
import re

# 读取第三个 Excel 文件，用于查找 IP 数据和 Service 数据
df3 = pd.read_excel("UAT_FW_Policy_IP.xlsx")

# 打印列名以检查实际的列名
print("UAT_FW_Policy_IP.xlsx:")
print(df3.columns.tolist())

# 创建一个字典，用于存储 Dst Network Object 和对应的 IP 地址，并去除'号
ip_dict = {}
for index, row in df3.iterrows():
    # 假设第一列是 Dst Network Object，第二列是 IP Address
    key = str(row.iloc[0]).replace("'", "")  # 去除'号
    value = str(row.iloc[1]).replace("'", "")  # 去除'号
    ip_dict[key] = value

# 读取第二个 Excel 文件，用于查找 Service 数据
df4 = pd.read_excel("UAT_FW_Services_object.xlsx")

# 打印列名以检查实际的列名
print("UAT_FW_Services_object.xlsx:")
print(df4.columns.tolist())

# 创建一个字典，用于存储 Service Name 和对应的 Protocol/Port 数据，并去除'号
service_dict = {}
for index, row in df4.iterrows():
    # 假设第一列是 #Service Name，第二列是 Protocol
    key = str(row.iloc[0]).replace("'", "")  # 去除'号
    protocol_port = str(row.iloc[2]).replace("'", "")  # 去除'号
    service_dict[key] = protocol_port

# 读取第一个 Excel 文件
df1 = pd.read_excel("UAT_FW_Policy_IPObject.xlsx")

# 提取需要的列，显式创建副本
selected_columns = ['#Name', ' Src Zone', ' Src Network Object', ' Dst Zone', ' Dst Network Object', ' Service']
df1_selected = df1[selected_columns].copy()  # 显式创建副本

# 重命名列以便合并
df1_selected.rename(columns={
    '#Name': 'Policy Name',
    ' Src Zone': 'Src Zone',
    ' Src Network Object': ' Src Network Object',
    ' Dst Zone': 'Dst Zone',
    ' Dst Network Object': ' Dst Network Object',
    ' Service': ' Service'
}, inplace=True)

# 读取第二个 Excel 文件
df2 = pd.read_excel("防火墙策略表-新加坡.xlsx")

# 合并两个 DataFrame，保留第二个文件的原有数据
df_merged = pd.concat([df2, df1_selected], ignore_index=True)

# 将逗号分隔的数据转换为换行符分隔
df_merged[' Src Network Object'] = df_merged[' Src Network Object'].apply(
    lambda x: x.replace(',', '\n') if isinstance(x, str) else x)
df_merged[' Dst Network Object'] = df_merged[' Dst Network Object'].apply(
    lambda x: x.replace(',', '\n') if isinstance(x, str) else x)

# 使用 openpyxl 读取原始工作簿以保留样式
wb = load_workbook("防火墙策略表-新加坡.xlsx")
ws = wb.active

# 设置字体样式
font = Font(size=11, color="000000")  # 11号黑色字体

# 写入合并后的数据并设置字体
for row_idx, row in df_merged.iterrows():
    for col_idx, value in enumerate(row, start=1):
        cell = ws.cell(row=row_idx + 2, column=col_idx, value=value)  # 从第2行开始写入数据
        cell.font = font  # 设置字体样式

# 查找对应的 IP 数据并写入 Source IP 列
for row_idx, row in df_merged.iterrows():
    src_network_objects = str(row[' Src Network Object']).split('\n')
    ip_addresses = []
    for obj in src_network_objects:
        cleaned_obj = obj.replace("'", "")  # 去除'号
        if cleaned_obj in ip_dict:
            # 将逗号分隔的 IP 地址转换为换行符分隔
            ip_address = ip_dict[cleaned_obj].replace(',', '\n')
            ip_addresses.append(ip_address)
    if ip_addresses:
        ws.cell(row=row_idx + 2, column=4, value='\n'.join(ip_addresses))  # 假设 Source IP 列是第4列

# 查找对应的 IP 数据并写入 Destination IP 列
for row_idx, row in df_merged.iterrows():
    dst_network_objects = str(row[' Dst Network Object']).split('\n')
    ip_addresses = []
    for obj in dst_network_objects:
        cleaned_obj = obj.replace("'", "")  # 去除'号
        if cleaned_obj in ip_dict:
            # 将逗号分隔的 IP 地址转换为换行符分隔
            ip_address = ip_dict[cleaned_obj].replace(',', '\n')
            ip_addresses.append(ip_address)
    if ip_addresses:
        ws.cell(row=row_idx + 2, column=7, value='\n'.join(ip_addresses))  # 假设 Destination IP 列是第7列

# 创建一个字典，用于存储 txt 文件中的 Protocol/Port 数据
txt_dict = {}
try:
    with open("tcpudp port(1).txt", "r") as file:
        config_text = file.read()  # 读取整个文件内容

        # 使用正则表达式匹配多种格式
        patterns = [
            (r'^(\S+)\s*\n(tcp|udp)/(\d+)\s*\n', re.MULTILINE | re.IGNORECASE),  # 格式1：service_name\nTCP/UDP/port
            (r'^(\S+)\s*\n.*?destination Port:(\d+)\b', re.MULTILINE | re.DOTALL | re.IGNORECASE),
            # 格式2：service_name destination Port: port
            (
            r'^(\S+)\s*\n((?:TCP|UDP): source Port:\d+-\d+ destination Port:(\d+))\s*\n', re.MULTILINE | re.IGNORECASE),
            # 格式3：service_name\nTCP: source Port:0-65535 destination Port:389
            (r'^(\S+)\s*\n.*?Port\s*=\s*(\d+)\b', re.MULTILINE | re.DOTALL | re.IGNORECASE)
            # 格式4：service_name Port = port
        ]

        for pattern in patterns:
            matches = re.finditer(pattern[0], config_text, pattern[1])
            for match in matches:
                service_name = match.group(1).strip().upper()  # 统一转换为大写
                if pattern[0] == patterns[0][0]:
                    proto = match.group(2).lower()
                    port = match.group(3)
                elif pattern[0] == patterns[1][0]:
                    proto = 'tcp'  # 默认协议为 TCP
                    port = match.group(2)
                elif pattern[0] == patterns[2][0]:
                    proto = match.group(2).split(':')[0].lower()  # TCP 或 UDP
                    port = match.group(3)
                elif pattern[0] == patterns[3][0]:
                    proto = 'tcp'  # 默认协议为 TCP
                    port = match.group(2)
                else:
                    proto = 'unknown'
                    port = 'unknown'

                # 如果服务名称已存在，追加协议/端口，否则新建列表
                if service_name in txt_dict:
                    txt_dict[service_name].append(f"{proto}/{port}")
                else:
                    txt_dict[service_name] = [f"{proto}/{port}"]

                print(f"Loaded from txt: {service_name} -> {proto}/{port}")

except Exception as e:
    print(f"Error reading txt file: {e}")

# 查找对应的 Service 数据并写入 Protocol 和 Service Port 列
for row_idx, row in df_merged.iterrows():
    services = str(row[' Service']).split(',')  # 以逗号分隔 Service 列的数据
    protocols = []
    service_ports = []
    for service in services:
        service = service.strip()
        print(f"\nProcessing service: {service}")

        if service.startswith('Predefined Service/'):
            # 提取 service 名称，去除 "Predefined Service/" 前缀
            keyword = service.split('/', 1)[1]
            keyword_cleaned = keyword.replace("'", "").strip().upper()  # 统一转换为大写
            print(f"  Extracted keyword: {keyword_cleaned}")

            # 查找匹配的 Protocol/Port
            if keyword_cleaned in txt_dict:
                for protocol_port in txt_dict[keyword_cleaned]:
                    print(f"  Matched service: {keyword_cleaned} -> {protocol_port}")
                    # 分离 Protocol 和 Port
                    if '/' in protocol_port:
                        proto, port = protocol_port.split('/', 1)
                        protocols.append(proto)
                        service_ports.append(port)
                    else:
                        protocols.append(protocol_port)
                        service_ports.append('')
            else:
                print(f"  No match found for service: {keyword_cleaned}")
                # 如果没有匹配到，写入原始服务名称
                protocols.append(service)
                service_ports.append('')
        elif service.startswith('Custom Services/'):
            # 提取 / 后面的部分
            keyword = service.split('/', 1)[1]
            # 去除单引号
            keyword_cleaned = keyword.replace("'", "")
            print(f"  Extracted keyword: {keyword_cleaned}")

            # 查找匹配的 Service
            if keyword_cleaned in service_dict:
                # 获取 Protocol/Port 数据
                protocol_port_data = service_dict[keyword_cleaned]
                print(f"  Matched service: {keyword_cleaned} -> {protocol_port_data}")
                # 以换行符分隔多组数据
                protocol_port_groups = protocol_port_data.split('\n')
                for group in protocol_port_groups:
                    # 以冒号分隔 Protocol 和 Port
                    if ':' in group:
                        protocol, port = group.split(':', 1)
                        protocol = protocol.strip()
                        port = port.strip()
                        # 将逗号分隔的 Port 转换为换行符分隔
                        port = port.replace(',', '\n')
                        protocols.append(protocol)
                        service_ports.append(port)
                        print(f"  Custom Service '{service}' matched. Protocol: {protocol}, Service Port: {port}")
                    else:
                        # 如果没有冒号，直接写入 Protocol 列，Service Port 列留空
                        protocols.append(group.strip())
                        service_ports.append('')
                        print(f"  Custom Service '{service}' matched. Protocol: {group.strip()}, Service Port: (None)")
            else:
                # 如果没有匹配到，写入提取的关键字到 Protocol 列，Service Port 列留空
                protocols.append(keyword_cleaned)
                service_ports.append('')
                print(f"  Custom Service '{service}' not matched. Protocol: {keyword_cleaned}, Service Port: (None)")
        else:
            # 如果不是 Custom Services，直接写入原始 Service 到 Protocol 列，Service Port 列留空
            protocols.append(service)
            service_ports.append('')
            print(f"  Service '{service}' not matched. Protocol: {service}, Service Port: (None)")

    # 对 Protocol 列进行去重（不区分大小写）
    if protocols:
        # 创建一个字典用于去重，键为小写的协议名称，值为原始协议名称
        unique_protocols_dict = {}
        for proto in protocols:
            lower_proto = proto.lower()
            if lower_proto not in unique_protocols_dict:
                unique_protocols_dict[lower_proto] = proto
        unique_protocols = list(unique_protocols_dict.values())
        print(f"  Unique protocols after deduplication: {unique_protocols}")
        ws.cell(row=row_idx + 2, column=8, value='\n'.join(unique_protocols))  # 假设 Protocol 列是第8列

    # 去重并写入 Service Port 列
    if service_ports:
        unique_service_ports = list(dict.fromkeys(service_ports))  # 使用 dict.fromkeys() 去重并保留顺序
        ws.cell(row=row_idx + 2, column=9, value='\n'.join(unique_service_ports))  # 假设 Service Port 列是第9列


for row in ws.iter_rows():
    # 获取行中的单元格列表
    cells = list(row)
    if len(cells) >=10:
        for cell in cells[9:]:
            cell.value = None

# 保存工作簿
wb.save("防火墙策略表-新加坡_updated.xlsx")

print("所有行的数据已写入完成！")